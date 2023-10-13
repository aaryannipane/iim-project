from flask import Flask, render_template, request, make_response, redirect, url_for, session
from openpyxl import Workbook, load_workbook
import pandas as pd
import sqlite3
import uuid
from product_attribute_perceptions import getAnalysis


app = Flask(__name__)
app.secret_key = 'my-secret-keys'
app.debug = True


# helper functions
def getUsers(username):
    connection = sqlite3.connect("user.db")
    cursor = connection.cursor()
    user = session.get('user')
    cursor.execute("SELECT * FROM USERS WHERE username != ?", (username,))

    users = cursor.fetchall()
    return users

@app.route('/', methods=['GET'])
def index():

    if session.get("message"):
        session.pop("message")

    # check user logged 
    # get cookie and find user 
    user = session.get("user")

    if not user:
        return redirect("/signin")
    
    return render_template('index.html')


@app.route('/save_media_allocation', methods=['POST'])
def save_media_allocation():
    if request.method == 'POST':
        # Extract the form data
        product = request.form.get('product')
        soc_media = request.form.get('soc_media')
        radio = request.form.get('radio')
        journals = request.form.get('journals')
        dmail = request.form.get('dmail')
        search = request.form.get('search')


        # Save the form data to the Excel file
        try:
            wb = load_workbook('i_media_allocation_adv_t.xlsx')
            ws = wb.active
        except FileNotFoundError:
            print("new file")
            wb = Workbook()
            ws = wb.active
            header_row = [
                'Advertisement Media', 'Baller', 'Banness', 'Caster', 'Camst', 'Danil', 'Daibi', 'AP', 'Faldo', 'Fano', 'Gats', 'Gamer']
            ws.append(header_row)

            #insert cols in sheet 
            header_col = ['SocMedia (%)', 'Radio (%)', 'Journals (%)', 'Dmail (%)', 'Search (%)']
            i = 2
            
            for col in header_col:
                ws['A'+str(i)] = header_col[i-2]
                i += 1
            wb.save('i_media_allocation_adv_t.xlsx')

        # append data col wise 
        colData = [soc_media, radio, journals, dmail, search]

        col_got = False
        i = 0
        for col in ws.iter_cols(min_col=2, min_row=1, max_col=12, max_row=6):
            # print(col)
            for cell in col:
                print(cell.internal_value == product)
                if cell.internal_value == product:
                    col_got=True
                    # print(col)
                    continue
                
                if col_got:
                    cell.value = colData[i]
                    i+=1

            if col_got:
                break

        
        wb.save('i_media_allocation_adv_t.xlsx')

        # Return the form template to display it again
        return render_template('index.html')
    else:
        # For GET request, simply render the form template
        return render_template('index.html')

# 
@app.route('/save_decision_summary', methods=['POST'])
def save_decision_summary():

    if request.method == 'POST':
        # Extract the form data
        company = request.form.get('company')
        product = request.form.get('product')
        research_project = request.form.get('research_project')
        market = request.form.get('market')
        creative_desk = request.form.get('creative_desk')
        advertising = request.form.get('advertising')
        price = request.form.get('price')
        production = request.form.get('production')
        disposal = request.form.get('disposal')

        # get year
        connection = sqlite3.connect("user.db")
        cursor = connection.cursor()
        cursor.execute("SELECT year FROM yearTable WHERE id = 1")

        data = cursor.fetchall()
        year = str(data[0][0])

        

        try:
            wb = load_workbook(f'./decision/i_Decision Summary_adv_t{year}.xlsx')
            ws = wb.active

        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            header_row = ['Company', 'Product', 'Research Project', 'Market', 'Creative Desk', 'Advertising', 'Price',
                          'Production', 'Disposal']
            ws.append(header_row)

        row = [company, product, research_project, market,
               creative_desk, advertising, price, production, disposal]
        ws.append(row)
        wb.save(f'./decision/i_Decision Summary_adv_t{year}.xlsx')

        # Return the form template to display it again
        return render_template('index.html')
    else:
        # For GET request, simply render the form template
        return render_template('index.html')

# 
@app.route('/save_shift', methods=['POST'])
def save_shift():
    if request.method == 'POST':
        # Extract the form data
        company = request.form.get('company')
        product = request.form.get('product')
        price = request.form.get('price')
        weight = request.form.get('weight')
        complexity = request.form.get('complexity')
        frequency = request.form.get('frequency')
        power = request.form.get('power')
        speed = request.form.get('speed')

         # get year
        connection = sqlite3.connect("user.db")
        cursor = connection.cursor()
        cursor.execute("SELECT year FROM yearTable WHERE id = 1")

        data = cursor.fetchall()
        year = str(data[0][0])
        print(str(data[0][0]))

        # Save the form data to the Excel file
        try:
            wb = load_workbook(f'./shift/i_shift_t{year}.xlsx')
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            header_row = ['Company', 'Product', 'Price', 'Weight',
                          'Complexity', 'Frequency', 'Power', 'Speed']
            ws.append(header_row)

        row = [company, product, price, weight,
               complexity, frequency, power, speed]
        ws.append(row)
        wb.save(f'./shift/i_shift_t{year}.xlsx')

        # Return the form template to display it again
        return render_template('index.html')
    else:
        # For GET request, simply render the form template
        return render_template('index.html')

@app.route('/save_product_attribute', methods=['POST'])
def save_product_attribute():
    if request.method == 'POST':
        # Extract the form data
        company = request.form.get('company')
        product = request.form.get('product')
        price = request.form.get('price')
        weight = request.form.get('weight')
        complexity = request.form.get('complexity')
        frequency = request.form.get('frequency')
        power = request.form.get('power')
        speed = request.form.get('speed')

        # Save the form data to the Excel file
        try:
            wb = load_workbook('c_Product_Attribute_Perceptions_only.xlsx')
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            header_row = ['Company', 'Product', 'Price', 'Weight',
                          'Complexity', 'Frequency', 'Power', 'Speed']
            ws.append(header_row)

        row = [company, product, price, weight,
               complexity, frequency, power, speed]
        ws.append(row)
        wb.save('c_Product_Attribute_Perceptions_only.xlsx')

        # Return the form template to display it again
        return render_template('index.html')
    else:
        # For GET request, simply render the form template
        return render_template('index.html')


@app.route('/result', methods=['GET'])
def result():
    # check if admin had runed the analysis
    # if done then only show result to student
    connection = sqlite3.connect("user.db")
    cursor = connection.cursor()

    cursor.execute("SELECT allowAccess FROM YEARTABLE WHERE id = 1")
    allow_access = cursor.fetchall()[0][0]

    if allow_access:
        df = pd.read_excel("./calculated_tables_t.xlsx", sheet_name="c_Product Attribute Perceptions")
        table_html = df.to_html(classes="table table-bordered table-striped table-hover mt-3", index=False)
        return render_template('result.html', table_html=table_html)
    else:
        return render_template('result.html', message="no result")

    return render_template('result.html')

@app.route("/show_result_access", methods=['GET'])
def show_result_access():
    connection = sqlite3.connect("user.db")
    cursor = connection.cursor()

    cursor.execute("SELECT allowAccess FROM YEARTABLE WHERE id = 1")
    allow_access = cursor.fetchall()[0][0]
    if allow_access == 0:
        allow_access = 1
    else:
        allow_access = 0
        
    cursor.execute("UPDATE YEARTABLE SET allowAccess = ? WHERE id = 1", (allow_access,))
    connection.commit()
    session['showResult'] = allow_access
    print(f'--> {allow_access}')
    return redirect("/admin")
    

@app.route('/run_r_script', methods=['GET'])
def run_r_script():
    # subprocess.call(['python', 'product_attribute_perceptions.py'])

    # if admin runs then increment year and return prevoius files analaysis
    # TODO: student will see analysed data after admin runs analysis 
    # increment analysed col of admin after analy
    if session.get('user')['username'] == "admin":
        # increment year 
        # return prevoius files analysis
        connection = sqlite3.connect("user.db")
        cursor = connection.cursor()

        # get year 
        cursor.execute("SELECT year FROM YEARTABLE WHERE id = 1")
        year = cursor.fetchall()[0][0]

        # increment year TODO: don't upgrade year after every click on analysis of admin
        cursor.execute("UPDATE yearTable SET year = ? WHERE id = 1", (year+1,))
        connection.commit()

        # try:
        getAnalysis(year)
        # except FileNotFoundError:
        # clear session after 
        #     print("file not found")
        #     session['error'] = "file not found"
        #     return redirect(url_for("admin_dashboard"))

        # analysis data to diff file with year-1 number
        df = pd.read_excel("./calculated_tables_t.xlsx", sheet_name="c_Product Attribute Perceptions")
        df_copy = df.copy()
        df_copy.to_excel(f'./product/c_Product_Attribute_Perceptions{year}.xlsx', index=False)

        html_table = df.to_html(classes="table table-bordered table-striped table-hover mt-3", index=False)
        return render_template("admin_dashboard.html", html_table=html_table)
    

    # get year from DB
    # check current year is less than admins analysedYear then only show data
    year = 1
    getAnalysis(1)

    # read file
    df = pd.read_excel("./calculated_tables_t.xlsx", sheet_name="c_Product Attribute Perceptions")
    
    table_html = df.to_html(classes="table table-bordered table-striped table-hover mt-3", index=False)

    return render_template('index.html', table_html=table_html)



# LOGIN / SIGN UP ROUTES
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        user_name = request.form.get("username")
        user_password = request.form.get("userpassword")
        user_type = request.form.get("userType")

        # db connection
        connection = sqlite3.connect("user.db")
        cursor = connection.cursor()

        # check same email exist in DB
        query = "SELECT username FROM USERS WHERE username='"+user_name+"'"
        cursor.execute(query)
        result = cursor.fetchall()
        if result:
           session['message'] = {'page':"signup",'type': "danger", 'message': "User Already Exists with that username"}
           return redirect("/signup")
        
        
        # if not exist insert user in db
        unique_id = str(uuid.uuid4())
        cursor.execute("INSERT INTO USERS (id, username, password, userType) VALUES (?, ?, ?, ?)", (unique_id, user_name, user_password, user_type))

        connection.commit()
        
        
        # redirect to login page
        if cursor.rowcount > 0:
            connection.close()
            if session.get("message"):
                session.pop("message")

            return redirect(url_for("admin_dashboard"))
        

        connection.close()
        return render_template("register.html", message="something went wrong, please try again later")

    if session.get("user") and session.get("user")["userType"] == "admin":
        return render_template("register.html")
    
    return redirect(url_for("signin"))

@app.route('/signin', methods=['GET', 'POST'])
def signin():
    if request.method == "POST":
        username = request.form.get("username")
        user_password = request.form.get("userpassword")

        # check user exists with that username and password
        connection = sqlite3.connect("user.db")
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM USERS WHERE username=?", (username,))
        result = cursor.fetchall()


        # if not exists then go to signin page
        if len(result) == 0:
            session['message'] = {'page':"signin",'type': "danger", 'message': "user does'nt exist"}
            return redirect("/signin")

        # if password not matched 
        if user_password != result[0][2]:
            session['message'] = {'page':"signin",'type': "danger", 'message': "invalid password"}
            return redirect("/signin")


        session['user'] = {'id':result[0][0], 'username':username, 'userType': result[0][3]}
        if session.get("message"):
            session.pop("message")

        # check type and route accordingly
        if result[0][3] == "admin":
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for("index"))
    
    return render_template("login.html")

@app.route('/logout', methods=['POST'])
def logout():
    if request.method == "POST":
        if not session.get("user"):
            return redirect("/signin")
        session.clear()
        return redirect("/signin")

# admin routes
@app.route("/admin", methods=["GET", "POST"])
def admin_dashboard():
    # check if session type is set to admin then only render admin else render student
    if not session.get('user'):
        return redirect(url_for('signin'))

    if session.get('user')['userType'] != "admin":
        return redirect(url_for('signin'))

    # get all users data and pass that data to admin in table format
    
    users = getUsers(session.get("user")['username'])

    # set users_table session
    session["users_table"] = users

    connection = sqlite3.connect("user.db")
    cursor = connection.cursor()

    cursor.execute("SELECT allowAccess FROM YEARTABLE WHERE id = 1")
    allow_access = cursor.fetchall()[0][0]
    
    session['showResult'] = allow_access

    return render_template("admin_dashboard.html")

# get data from files
@app.route("/get/file", methods=['POST'])
def getFile():
    if request.method == "POST":
        fileName = request.form.get('action')
        
        # fetch year from database then load file
        year = "1"

        classes = 'table table-bordered table-striped table-hover mt-3'
        if fileName == "decision":
            df = pd.read_excel(f"./decision/i_Decision Summary_adv_t{year}.xlsx")
            table_html = df.to_html(classes=classes, index=False)
            session['file_table'] = table_html
            session['file_name'] = fileName
            return redirect(url_for("admin_dashboard"))
        elif fileName == "shift":
            df = pd.read_excel(f"./shift/i_shift_t{year}.xlsx")
            table_html = df.to_html(classes=classes, index=False)
            session['file_table'] = table_html
            session['file_name'] = fileName
            return redirect(url_for("admin_dashboard"))
        elif fileName == "media":
            df = pd.read_excel(f"./media/i_media_allocation_adv_t.xlsx")
            table_html = df.to_html(classes=classes, index=False)
            session['file_table'] = table_html
            session['file_name'] = fileName
            return redirect(url_for("admin_dashboard"))
    

    return redirect(url_for("admin_dashboard"))

@app.route("/reset", methods=["GET"])
def reset():
    connection = sqlite3.connect("user.db")
    cursor = connection.cursor()
    cursor.execute("UPDATE yearTable SET year = 1 WHERE id = 1")
    connection.commit()
    connection.close()

    return redirect(url_for("admin_dashboard"))


if __name__ == '__main__':
    app.run()
